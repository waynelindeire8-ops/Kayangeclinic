import json, os, uuid
from sqlite3 import IntegrityError
from flask import Blueprint, request, jsonify, render_template, send_from_directory
from flask_jwt_extended import get_jwt
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from config import Config
from app.database import get_db, _recreate_patients_table
from app.auth import login_required, log_audit

patients_bp = Blueprint('patients', __name__, url_prefix='/patients')


@patients_bp.route('/', strict_slashes=False)
@login_required
def list_page():
    db = get_db()
    providers = db.execute('SELECT id, name FROM insurance_providers ORDER BY name').fetchall()
    db.close()
    providers_json = json.dumps({p['id']: p['name'] for p in providers})
    return render_template('patients/list.html', providers_json=providers_json)


@patients_bp.route('/new', strict_slashes=False)
@login_required
def new_page():
    return render_template('patients/form.html')


@patients_bp.route('/<int:id>', strict_slashes=False)
@login_required
def detail_page(id):
    db = get_db()
    providers = db.execute('SELECT id, name FROM insurance_providers ORDER BY name').fetchall()
    db.close()
    providers_json = json.dumps({p['id']: p['name'] for p in providers})
    return render_template('patients/detail.html', patient_id=id, providers_json=providers_json)


@patients_bp.route('/<int:id>/edit', strict_slashes=False)
@login_required
def edit_page(id):
    return render_template('patients/form.html', patient_id=id)


@patients_bp.route('/api', methods=['GET'])
@login_required
def api_list():
    db = get_db()
    search = request.args.get('search', '')
    if search:
        patients = db.execute(
            '''SELECT * FROM patients
               WHERE first_name LIKE ? OR last_name LIKE ? OR patient_id LIKE ? OR phone LIKE ?
               ORDER BY created_at DESC''',
            (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%')
        ).fetchall()
    else:
        patients = db.execute('SELECT * FROM patients ORDER BY created_at DESC').fetchall()
    db.close()
    return jsonify([dict(p) for p in patients])


@patients_bp.route('/api/<int:id>', methods=['GET'])
@login_required
def api_get(id):
    db = get_db()
    patient = db.execute(
        '''SELECT p.*, ip.name as scheme_name 
           FROM patients p 
           LEFT JOIN insurance_providers ip ON p.scheme_id = ip.id
           WHERE p.id = ?''', (id,)).fetchone()
    if not patient:
        db.close()
        return jsonify({'error': 'Patient not found'}), 404

    allergies = db.execute('SELECT * FROM patient_allergies WHERE patient_id = ?', (id,)).fetchall()
    medical_history = db.execute('SELECT * FROM patient_medical_history WHERE patient_id = ?', (id,)).fetchall()
    medications = db.execute('SELECT * FROM patient_medications WHERE patient_id = ?', (id,)).fetchall()
    db.close()

    result = dict(patient)
    result['allergies'] = [dict(a) for a in allergies]
    result['medical_history'] = [dict(m) for m in medical_history]
    result['medications'] = [dict(m) for m in medications]
    return jsonify(result)


@patients_bp.route('/api', methods=['POST'])
@login_required
def api_create():
    current_user = get_jwt()
    data = request.json

    db = get_db()
    last = db.execute('SELECT patient_id FROM patients ORDER BY id DESC LIMIT 1').fetchone()
    if last:
        num = int(last['patient_id'].replace('KMC-', '')) + 1
    else:
        num = 1001
    patient_id = f'KMC-{num}'

    cursor = db.execute(
        '''INSERT INTO patients (patient_id, first_name, last_name, dob, gender, phone, email, address,
           emergency_contact_name, emergency_contact_phone, blood_group, scheme_provider, scheme_type, scheme_id, scheme_number)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (patient_id, data['first_name'], data['last_name'], data['dob'], data.get('gender'),
         data['phone'], data.get('email'), data.get('address'), data.get('emergency_contact_name'),
         data.get('emergency_contact_phone'), data.get('blood_group'), data.get('scheme_provider'),
         data.get('scheme_type'), data.get('scheme_id'), data.get('scheme_number'))
    )
    new_id = cursor.lastrowid

    for allergy in data.get('allergies', []):
        db.execute('INSERT INTO patient_allergies (patient_id, allergy) VALUES (?, ?)', (new_id, allergy))

    for condition in data.get('medical_history', []):
        db.execute('INSERT INTO patient_medical_history (patient_id, condition_name, diagnosis_date, notes) VALUES (?, ?, ?, ?)',
                   (new_id, condition.get('condition_name'), condition.get('diagnosis_date'), condition.get('notes')))

    for med in data.get('medications', []):
        db.execute('INSERT INTO patient_medications (patient_id, medication_name, dosage, frequency, prescribed_by, start_date, end_date) VALUES (?, ?, ?, ?, ?, ?, ?)',
                   (new_id, med.get('medication_name'), med.get('dosage'), med.get('frequency'),
                    med.get('prescribed_by'), med.get('start_date'), med.get('end_date')))

    db.commit()
    log_audit(current_user['id'], current_user['username'], 'create', 'patient', new_id,
              f'Created patient {data["first_name"]} {data["last_name"]}', request.remote_addr)
    db.close()

    return jsonify({'id': new_id, 'patient_id': patient_id}), 201


@patients_bp.route('/api/<int:id>', methods=['PUT'])
@login_required
def api_update(id):
    current_user = get_jwt()
    data = request.json

    db = get_db()
    db.execute(
        '''UPDATE patients SET first_name=?, last_name=?, dob=?, gender=?, phone=?, email=?, address=?,
           emergency_contact_name=?, emergency_contact_phone=?, blood_group=?, scheme_provider=?, scheme_type=?,
           scheme_id=?, scheme_number=?, updated_at=CURRENT_TIMESTAMP WHERE id=?''',
        (data['first_name'], data['last_name'], data['dob'], data.get('gender'),
         data['phone'], data.get('email'), data.get('address'), data.get('emergency_contact_name'),
         data.get('emergency_contact_phone'), data.get('blood_group'), data.get('scheme_provider'),
         data.get('scheme_type'), data.get('scheme_id'), data.get('scheme_number'), id)
    )

    db.execute('DELETE FROM patient_allergies WHERE patient_id = ?', (id,))
    for allergy in data.get('allergies', []):
        db.execute('INSERT INTO patient_allergies (patient_id, allergy) VALUES (?, ?)', (id, allergy))

    db.execute('DELETE FROM patient_medical_history WHERE patient_id = ?', (id,))
    for condition in data.get('medical_history', []):
        db.execute('INSERT INTO patient_medical_history (patient_id, condition_name, diagnosis_date, notes) VALUES (?, ?, ?, ?)',
                   (id, condition.get('condition_name'), condition.get('diagnosis_date'), condition.get('notes')))

    db.execute('DELETE FROM patient_medications WHERE patient_id = ?', (id,))
    for med in data.get('medications', []):
        db.execute('INSERT INTO patient_medications (patient_id, medication_name, dosage, frequency, prescribed_by, start_date, end_date) VALUES (?, ?, ?, ?, ?, ?, ?)',
                   (id, med.get('medication_name'), med.get('dosage'), med.get('frequency'),
                    med.get('prescribed_by'), med.get('start_date'), med.get('end_date')))

    db.commit()
    log_audit(current_user['id'], current_user['username'], 'update', 'patient', id,
              f'Updated patient {data["first_name"]} {data["last_name"]}', request.remote_addr)
    db.close()

    return jsonify({'message': 'Patient updated successfully'})


@patients_bp.route('/api/<int:id>', methods=['DELETE'])
@login_required
def api_delete(id):
    current_user = get_jwt()
    db = get_db()
    patient = db.execute('SELECT * FROM patients WHERE id = ?', (id,)).fetchone()
    if not patient:
        db.close()
        return jsonify({'error': 'Patient not found'}), 404
    try:
        db.execute('DELETE FROM patients WHERE id = ?', (id,))
        db.commit()
    except IntegrityError:
        db.close()
        return jsonify({'error': 'Cannot delete patient with existing consultations, lab tests, billing records, or other related data. Deactivate the patient instead.'}), 409
    log_audit(current_user['id'], current_user['username'], 'delete', 'patient', id,
              f'Deleted patient {patient["first_name"]} {patient["last_name"]}', request.remote_addr)
    db.close()
    return jsonify({'message': 'Patient deleted successfully'})


@patients_bp.route('/api/<int:id>/history', methods=['GET'])
@login_required
def api_history(id):
    db = get_db()
    consultations = db.execute(
        '''SELECT c.*, u.first_name as doctor_first, u.last_name as doctor_last
           FROM consultations c LEFT JOIN users u ON c.doctor_id = u.id
           WHERE c.patient_id = ? ORDER BY c.created_at DESC''', (id,)).fetchall()
    lab_tests = db.execute(
        '''SELECT l.*, u.first_name as doctor_first, u.last_name as doctor_last
           FROM lab_tests l LEFT JOIN users u ON l.doctor_id = u.id
           WHERE l.patient_id = ? ORDER BY l.created_at DESC''', (id,)).fetchall()
    procedures = db.execute(
        '''SELECT p.*, u.first_name as doctor_first, u.last_name as doctor_last
           FROM procedures p LEFT JOIN users u ON p.performed_by = u.id
           WHERE p.patient_id = ? ORDER BY p.created_at DESC''', (id,)).fetchall()
    billing = db.execute(
        'SELECT * FROM billing WHERE patient_id = ? ORDER BY created_at DESC', (id,)).fetchall()
    db.close()
    return jsonify({
        'consultations': [dict(c) for c in consultations],
        'lab_tests': [dict(l) for l in lab_tests],
        'procedures': [dict(p) for p in procedures],
        'billing': [dict(b) for b in billing]
    })


@patients_bp.route('/import', strict_slashes=False)
@login_required
def import_page():
    db = get_db()
    providers = db.execute('SELECT id, name FROM insurance_providers ORDER BY name').fetchall()
    db.close()
    return render_template('patients/import.html', providers=providers)


@patients_bp.route('/api/import', methods=['POST'])
@login_required
def api_import():
    try:
        return _api_import()
    except Exception as e:
        return jsonify({'error': f'Import error: {str(e)}'}), 500


def _api_import():
    current_user = get_jwt()
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400

    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

    field_map = {
        'first_name': ['first name', 'first_name', 'firstname', 'given name', 'fname'],
        'last_name': ['last name', 'last_name', 'lastname', 'surname', 'family name', 'lname'],
        'dob': ['dob', 'date of birth', 'birth date', 'birthdate', 'birthday'],
        'gender': ['gender', 'sex'],
        'phone': ['phone', 'phone number', 'mobile', 'cell', 'telephone'],
        'email': ['email', 'email address'],
        'address': ['address', 'home address', 'street address'],
        'emergency_contact_name': ['emergency contact', 'emergency contact name', 'next of kin', 'emergency name'],
        'emergency_contact_phone': ['emergency phone', 'emergency contact phone', 'next of kin phone', 'emergency phone number'],
        'blood_group': ['blood group', 'blood type', 'bloodgroup'],
        'scheme_provider': ['insurance', 'scheme provider', 'insurance provider', 'provider', 'scheme'],
        'scheme_type': ['scheme type', 'plan', 'insurance type', 'cover type', 'scheme'],
        'scheme_number': ['scheme number', 'policy number', 'member number', 'insurance number', 'policy no'],
    }

    col_mapping = {}
    for field, aliases in field_map.items():
        for i, header in enumerate(headers):
            if header in aliases:
                col_mapping[field] = i
                break

    if 'first_name' not in col_mapping or 'last_name' not in col_mapping:
        return jsonify({'error': 'Excel must have "First Name" and "Last Name" columns. Found: ' + ', '.join(headers)}), 400

    db = get_db()
    db.execute("PRAGMA foreign_keys = OFF")

    # Ensure dob and phone are nullable
    for r in db.execute("PRAGMA table_info(patients)").fetchall():
        if r[1] in ('dob', 'phone') and r[3] == 1:
            _recreate_patients_table(db)
            break

    # Build insurance provider lookup (name -> id, case-insensitive)
    providers_raw = db.execute('SELECT id, name FROM insurance_providers').fetchall()
    provider_map = {}
    for p in providers_raw:
        provider_map[p['name'].strip().lower()] = p['id']

    last = db.execute('SELECT patient_id FROM patients ORDER BY id DESC LIMIT 1').fetchone()
    if last:
        num = int(last['patient_id'].replace('KMC-', ''))
    else:
        num = 1000

    created = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        first_name = str(row[col_mapping['first_name']] or '').strip()
        last_name = str(row[col_mapping['last_name']] or '').strip()
        if not first_name or not last_name:
            skipped += 1
            continue

        num += 1
        patient_id = f'KMC-{num}'

        def get_val(field):
            if field in col_mapping:
                val = row[col_mapping[field]]
                if val is None:
                    return None
                if hasattr(val, 'strftime'):
                    return val.strftime('%Y-%m-%d')
                return str(val).strip()
            return None

        dob = get_val('dob')
        if dob and len(dob) == 10 and dob[4] == '-' and dob[7] == '-':
            pass
        elif dob and '/' in dob:
            parts = dob.split('/')
            if len(parts) == 3:
                dob = f'{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}'
        else:
            dob = None

        gender = get_val('gender')
        if gender:
            gender = gender.capitalize()
            if gender not in ('Male', 'Female', 'Other'):
                gender = None

        scheme_provider_val = get_val('scheme_provider')
        scheme_id_val = None
        if scheme_provider_val:
            key = scheme_provider_val.strip().lower()
            if key in provider_map:
                scheme_id_val = provider_map[key]
            else:
                matches = []
                for pname, pid in provider_map.items():
                    if pname.startswith(key) or key.startswith(pname):
                        matches.append((pname, pid))
                if len(matches) == 1:
                    scheme_id_val = matches[0][1]

        try:
            cursor = db.execute(
                '''INSERT INTO patients (patient_id, first_name, last_name, dob, gender, phone, email, address,
                   emergency_contact_name, emergency_contact_phone, blood_group, scheme_provider, scheme_type, scheme_id, scheme_number)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (patient_id, first_name, last_name, dob or None, gender,
                 get_val('phone') or None, get_val('email'), get_val('address'),
                 get_val('emergency_contact_name'), get_val('emergency_contact_phone'),
                 get_val('blood_group'), scheme_provider_val, get_val('scheme_type'),
                 scheme_id_val, get_val('scheme_number'))
            )
            created += 1
        except Exception as e:
            errors.append(f'Row {row_idx}: {str(e)}')
            skipped += 1

    db.commit()
    db.execute("PRAGMA foreign_keys = ON")
    log_audit(current_user['id'], current_user['username'], 'import', 'patients', None,
              f'Imported {created} patients from Excel', request.remote_addr)
    db.close()

    return jsonify({
        'message': f'Import complete: {created} created, {skipped} skipped',
        'created': created,
        'skipped': skipped,
        'errors': errors[:20]
    })


ALLOWED_MIMETYPES = {'application/pdf', 'image/jpeg', 'image/png', 'image/gif', 'image/webp'}


@patients_bp.route('/api/<int:patient_id>/documents', methods=['GET'])
@login_required
def api_documents_list(patient_id):
    db = get_db()
    docs = db.execute(
        '''SELECT pd.*, u.first_name || ' ' || u.last_name as uploaded_by_name
           FROM patient_documents pd
           LEFT JOIN users u ON pd.uploaded_by = u.id
           WHERE pd.patient_id = ?
           ORDER BY pd.created_at DESC''', (patient_id,)
    ).fetchall()
    db.close()
    return jsonify([dict(d) for d in docs])


@patients_bp.route('/api/<int:patient_id>/documents', methods=['POST'])
@login_required
def api_documents_upload(patient_id):
    current_user = get_jwt()
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.pdf', '.jpg', '.jpeg', '.png', '.gif', '.webp'):
        return jsonify({'error': 'Allowed: PDF, JPEG, PNG, GIF, WEBP'}), 400

    os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
    unique_name = f'{uuid.uuid4().hex}{ext}'
    filepath = os.path.join(Config.UPLOAD_FOLDER, unique_name)
    file.save(filepath)
    file_size = os.path.getsize(filepath)
    mime_type = file.content_type or 'application/octet-stream'
    notes = request.form.get('notes', '')

    db = get_db()
    cursor = db.execute(
        '''INSERT INTO patient_documents (patient_id, filename, original_filename, file_size, mime_type, notes, uploaded_by)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (patient_id, unique_name, file.filename, file_size, mime_type, notes, current_user['id'])
    )
    doc_id = cursor.lastrowid
    db.commit()
    db.close()
    log_audit(current_user['id'], current_user['username'], 'upload', 'patient_document', doc_id,
              f'Uploaded {file.filename} for patient {patient_id}', request.remote_addr)
    return jsonify({'id': doc_id, 'message': 'File uploaded'}), 201


@patients_bp.route('/api/documents/<int:doc_id>/download')
@login_required
def api_documents_download(doc_id):
    db = get_db()
    doc = db.execute('SELECT * FROM patient_documents WHERE id = ?', (doc_id,)).fetchone()
    db.close()
    if not doc:
        return jsonify({'error': 'Document not found'}), 404
    return send_from_directory(Config.UPLOAD_FOLDER, doc['filename'],
                               download_name=doc['original_filename'])


@patients_bp.route('/api/documents/<int:doc_id>', methods=['DELETE'])
@login_required
def api_documents_delete(doc_id):
    current_user = get_jwt()
    db = get_db()
    doc = db.execute('SELECT * FROM patient_documents WHERE id = ?', (doc_id,)).fetchone()
    if not doc:
        db.close()
        return jsonify({'error': 'Document not found'}), 404
    filepath = os.path.join(Config.UPLOAD_FOLDER, doc['filename'])
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
    except OSError:
        pass
    db.execute('DELETE FROM patient_documents WHERE id = ?', (doc_id,))
    db.commit()
    log_audit(current_user['id'], current_user['username'], 'delete', 'patient_document', doc_id,
              f'Deleted document {doc["original_filename"]}', request.remote_addr)
    db.close()
    return jsonify({'message': 'Document deleted'})
