import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Patients Template"

# Headers
headers = [
    'first_name', 'last_name', 'dob', 'gender', 'phone', 'email', 'address',
    'emergency_contact_name', 'emergency_contact_phone', 'blood_group',
    'scheme_provider', 'scheme_type', 'scheme_number'
]

# Sample data
sample_data = [
    ['John', 'Doe', '1990-01-15', 'Male', '0881234567', 'john.doe@email.com', '123 Main St, Blantyre',
     'Jane Doe', '0997654321', 'O+', 'NHIMA', 'Family', 'NHIMA-001234'],
    ['Mary', 'Smith', '1985-03-22', 'Female', '0999876543', 'mary.smith@email.com', '456 Oak Ave, Lilongwe',
     'Robert Smith', '0888765432', 'A+', 'MASM', 'Individual', 'MASM-567890'],
    ['James', 'Banda', '2000-07-10', 'Male', '0887777777', 'james.banda@email.com', '789 Pine Rd, Mzuzu',
     'Sarah Banda', '0999999999', 'B+', 'PVT', '', ''],
]

# Styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write sample data
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Set column widths
column_widths = {
    'A': 15, 'B': 15, 'C': 12, 'D': 10, 'E': 15, 'F': 25, 'G': 25,
    'H': 20, 'I': 15, 'J': 10, 'K': 15, 'L': 15, 'M': 20
}
for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Add instructions sheet
ws2 = wb.create_sheet("Instructions")
instructions = [
    "PATIENT IMPORT TEMPLATE INSTRUCTIONS",
    "",
    "1. Fill in patient data in the 'Patients Template' sheet",
    "2. Required columns: First Name, Last Name",
    "3. Date of Birth format: YYYY-MM-DD (e.g., 1990-01-15)",
    "4. Gender: Male, Female, or Other",
    "5. Phone: Malawi format (e.g., 0881234567, 0991234567)",
    "6. Blood Group: O+, O-, A+, A-, B+, B-, AB+, AB-",
    "7. Insurance Scheme: NHIMA, MASM, PVT, or other provider names",
    "8. Scheme Type: Family, Individual, Corporate, etc.",
    "9. Do NOT modify column headers",
    "10. Save as .xlsx format",
    "11. Upload via Patients > Import",
    "",
    "EXAMPLE VALUES:",
    "Gender: Male, Female, Other",
    "Blood Group: O+, O-, A+, A-, B+, B-, AB+, AB-",
    "Scheme Provider: NHIMA, MASM, PVT, First Capital, Axa, Old Mutual, Jubilee, NICO, RESMAID, MedHealth, COMAID, ESCOM, MRA, NABMAS",
    "Scheme Type: Family, Individual, Corporate, Executive, VIP, VVIP",
]

for i, line in enumerate(instructions, 1):
    ws2.cell(row=i, column=1, value=line)
    if line and not line.startswith(" ") and line.isupper():
        ws2.cell(row=i, column=1).font = Font(bold=True)

ws2.column_dimensions['A'].width = 80

wb.save('patient_import_template.xlsx')
print('Template created: patient_import_template.xlsx')